import openpyxl
import sys

from Anki import Anki
from Cleaner import Cleaner

SEPARATOR = ';'


def get_frequency(frequency_sheet, word: str) -> str:
    for i in range(4, frequency_sheet.max_row):
        if frequency_sheet.cell(row=i, column=2).value == word:
            # returns position on frequency list
            return str(frequency_sheet.cell(row=i, column=1).value)
    return "Out of range"


cleaner = Cleaner(sys.argv[1])
cleaner.cleanSubtitles()

wb = openpyxl.load_workbook('frequency.xlsx')
# print(wb.sheetnames)
frequencySheet = wb['10000k']
# print(sheet['A1'].value)


subtitles = open('subtitles', 'r')
to_anki = open('to_anki', 'w')

vocabulary = list(input("Podaj po przecinku wyrażenia jakie mam poszukać\n").split(','))
comment = input("Podaj komentarz jaki chcesz umiescic w kartach\n")

fileLineNumbers = 0
for line in subtitles:
    for vocab in vocabulary:
        if vocab in line:
            anki = Anki(vocab)
            pronunciation = anki.getPronunciation()
            polishMeanings = anki.getPolishMeanings()
            frequency = get_frequency(frequencySheet, vocab)
            line = line.replace("\n", "")
            # TODO paste all possible meaningsp
            to_anki.write(polishMeanings[0] + ';;' + vocab + ';' + line + ';' + comment + ';' + pronunciation[0] + ';' +
                          pronunciation[1] + ';' + frequency + '\n')
            # Only find first sentence with that vocabulary
            print(vocab, 'added.')
            vocabulary.remove(vocab)
            fileLineNumbers += 1
print(fileLineNumbers, 'rows added.')

if not vocabulary:
    print('Vocabulary which were not added')
    print(vocabulary)
else:
    print('All wocabulary were added')
